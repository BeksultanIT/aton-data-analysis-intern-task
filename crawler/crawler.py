import os
import csv
import zipfile
import sqlite3
import hashlib
import argparse
import tempfile
import logging
from pathlib import Path
from datetime import datetime
from typing import Optional

try:
    import docx
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False
    logging.warning("python-docx не установлен, .docx файлы будут пропущены")

try:
    import openpyxl
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False
    logging.warning("openpyxl не установлен, .xlsx файлы будут пропущены")

try:
    import pdfplumber
    HAS_PDF = True
except ImportError:
    HAS_PDF = False
    logging.warning("pdfplumber не установлен, .pdf файлы будут пропущены")

try:
    import py7zr
    HAS_7Z = True
except ImportError:
    HAS_7Z = False
    logging.warning("py7zr не установлен, .7z архивы будут пропущены")


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S"
)
log = logging.getLogger(__name__)


def parse_txt(path: str) -> str:
    """Читаем .txt с автодетектом кодировки."""
    for enc in ("utf-8", "cp1251", "latin-1"):
        try:
            with open(path, "r", encoding=enc) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
    return ""


def parse_docx(path: str) -> str:
    if not HAS_DOCX:
        return ""
    try:
        doc = docx.Document(path)
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    except Exception as e:
        log.warning(f"Ошибка парсинга docx {path}: {e}")
        return ""


def parse_xlsx(path: str) -> str:
    if not HAS_XLSX:
        return ""
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        parts = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                row_str = " | ".join(str(c) for c in row if c is not None)
                if row_str.strip():
                    parts.append(row_str)
        wb.close()
        return "\n".join(parts)
    except Exception as e:
        log.warning(f"Ошибка парсинга xlsx {path}: {e}")
        return ""


def parse_pdf(path: str) -> str:
    if not HAS_PDF:
        return ""
    try:
        with pdfplumber.open(path) as pdf:
            pages_text = []
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    pages_text.append(t)
        return "\n".join(pages_text)
    except Exception as e:
        log.warning(f"Ошибка парсинга pdf {path}: {e}")
        return ""


PARSERS = {
    ".txt":  parse_txt,
    ".docx": parse_docx,
    ".xlsx": parse_xlsx,
    ".xls":  parse_xlsx,
    ".pdf":  parse_pdf,
    ".csv":  parse_txt,
}


def extract_text(filepath: str) -> Optional[str]:
    ext = Path(filepath).suffix.lower()
    parser = PARSERS.get(ext)
    if parser is None:
        return None
    return parser(filepath)


def file_hash(path: str) -> str:
    h = hashlib.md5()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()



def unpack_zip(archive_path: str, dest_dir: str):
    with zipfile.ZipFile(archive_path, "r") as zf:
        zf.extractall(dest_dir)


def unpack_7z(archive_path: str, dest_dir: str):
    if not HAS_7Z:
        log.warning("py7zr не установлен, пропускаем .7z")
        return
    with py7zr.SevenZipFile(archive_path, mode="r") as sz:
        sz.extractall(path=dest_dir)


def process_archive(archive_path: str, parent_archive: str = "") -> list[dict]:
    records = []
    ext = Path(archive_path).suffix.lower()

    with tempfile.TemporaryDirectory() as tmpdir:
        try:
            if ext == ".zip":
                unpack_zip(archive_path, tmpdir)
            elif ext == ".7z":
                unpack_7z(archive_path, tmpdir)
            else:
                return records
        except Exception as e:
            log.warning(f"Не удалось распаковать {archive_path}: {e}")
            return records

        archive_label = parent_archive or archive_path
        for r in crawl_directory(tmpdir, archive_label=archive_label):
            r["source_archive"] = archive_label
            records.append(r)

    return records


ARCHIVE_EXTS = {".zip", ".7z", ".rar"}
SUPPORTED_EXTS = set(PARSERS.keys()) | ARCHIVE_EXTS


def crawl_directory(root: str, archive_label: str = "") -> list[dict]:
    records = []
    seen_hashes: set[str] = set()

    for dirpath, _, filenames in os.walk(root):
        for fname in filenames:
            fpath = os.path.join(dirpath, fname)
            ext = Path(fname).suffix.lower()

            if ext not in SUPPORTED_EXTS:
                continue

            if ext in ARCHIVE_EXTS:
                log.info(f"Архив: {fpath}")
                sub_records = process_archive(fpath, parent_archive=fpath)
                records.extend(sub_records)
                continue

            try:
                fhash = file_hash(fpath)
            except OSError:
                continue

            if fhash in seen_hashes:
                log.debug(f"Дубликат пропущен: {fpath}")
                continue
            seen_hashes.add(fhash)

            text = extract_text(fpath)
            if text is None:
                continue

            stat = os.stat(fpath)
            records.append({
                "file_path":      fpath,
                "file_name":      fname,
                "extension":      ext.lstrip("."),
                "size_bytes":     stat.st_size,
                "modified_at":    datetime.fromtimestamp(stat.st_mtime).isoformat(),
                "source_archive": archive_label,
                "content":        text.strip(),
                "file_hash":      fhash,
            })
            log.info(f"Обработан: {fname} ({len(text)} символов)")

    return records


FIELDNAMES = [
    "file_path", "file_name", "extension",
    "size_bytes", "modified_at", "source_archive",
    "content", "file_hash",
]


def save_csv(records: list[dict], output_path: str):
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
        writer.writeheader()
        writer.writerows(records)
    log.info(f"CSV сохранён: {output_path} ({len(records)} записей)")



def load_to_sqlite(records: list[dict], db_path: str):
    """
    Создаём две таблицы:
      - documents: метаданные файлов
      - documents_fts: виртуальная таблица FTS5 для полнотекстового поиска
    """
    os.makedirs(os.path.dirname(db_path) or ".", exist_ok=True)
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS documents (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            file_path     TEXT,
            file_name     TEXT,
            extension     TEXT,
            size_bytes    INTEGER,
            modified_at   TEXT,
            source_archive TEXT,
            content       TEXT,
            file_hash     TEXT UNIQUE
        )
    """)

    cur.execute("""
        CREATE VIRTUAL TABLE IF NOT EXISTS documents_fts USING fts5(
            file_name,
            content,
            content='documents',
            content_rowid='id'
        )
    """)

    inserted = 0
    for r in records:
        try:
            cur.execute("""
                INSERT OR IGNORE INTO documents
                    (file_path, file_name, extension, size_bytes,
                     modified_at, source_archive, content, file_hash)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                r["file_path"], r["file_name"], r["extension"],
                r["size_bytes"], r["modified_at"], r["source_archive"],
                r["content"], r["file_hash"],
            ))
            if cur.rowcount:
                inserted += 1
        except sqlite3.Error as e:
            log.warning(f"Ошибка вставки {r['file_name']}: {e}")

    cur.execute("INSERT INTO documents_fts(documents_fts) VALUES('rebuild')")

    conn.commit()
    conn.close()
    log.info(f"SQLite база: {db_path} (добавлено {inserted} документов)")



def main():
    parser = argparse.ArgumentParser(description="Краулер документов → CSV + SQLite FTS5")
    parser.add_argument("--root",   default="storage/", help="Корневая директория хранилища")
    parser.add_argument("--output", default="output/index.csv", help="Путь к итоговому CSV")
    parser.add_argument("--db",     default="output/fti.db",    help="Путь к SQLite базе")
    args = parser.parse_args()

    log.info(f"Старт краулинга: {args.root}")
    records = crawl_directory(args.root)
    log.info(f"Всего документов найдено: {len(records)}")

    if not records:
        log.warning("Нет документов для обработки, выходим")
        return

    save_csv(records, args.output)
    load_to_sqlite(records, args.db)
    log.info("Готово!")


if __name__ == "__main__":
    main()